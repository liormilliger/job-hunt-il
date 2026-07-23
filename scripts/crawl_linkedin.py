"""
LinkedIn Crawler — job-hunt-il
==================================================
Run this yourself when you want a LinkedIn Jobs scan.
A real Chrome window opens. You must be logged in to LinkedIn
(the script will wait for you to log in if needed).
The script scores every matching job and writes results to the Candidates tab.

Searches LinkedIn's Israel geo for the titles configured in config.json.

Run:  python crawl_linkedin_manual.py
"""

import sys
import os
import asyncio
import json
import re
import time
import datetime
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.async_api import async_playwright

from crawler_filters import should_skip

sys.stdout.reconfigure(encoding="utf-8")

# ── CONFIG ───────────────────────────────────────────────────────────────────
import jh_config
_CFG = jh_config.load()
TRACKER_PATH = _CFG["tracker"]

ANTHROPIC_API_KEY = jh_config.api_key(_CFG)

SCORE_THRESHOLD = 70
MAX_JOBS_PER_SEARCH = 25
SLEEP_BETWEEN_JOBS = 2.0

# LinkedIn searches come from config: each search title crossed with the
# configured geo (default 101620260 = Israel). (query, geoId, label) tuples.
_GEO = str(_CFG.get("linkedin_geo_id", "101620260"))
LINKEDIN_SEARCHES = [(t, _GEO, "Israel") for t in _CFG.get("search_titles", [])]

INTL_COMPANY_LIST = [
    "nike", "adidas", "l'oreal", "loreal", "henkel", "p&g", "procter", "unilever",
    "amazon", "dhl", "maersk", "ups", "fedex", "h&m", "ikea", "zara", "inditex",
    "nestle", "abbott", "johnson", "pfizer", "novartis", "roche", "siemens",
    "microsoft", "google", "apple", "intel", "applied materials", "kla", "nova",
    "elta", "rafael", "elbit", "checkpoint", "amdocs", "nice", "radcom",
    "teva", "given imaging", "orbotech", "strauss", "osem", "tnuva",
]
INTL_JD_KEYWORDS = [
    "emea", "global", "international", "cross-border", "import", "export",
    "fluent english", "english required", "english speaker", "multinational",
    "worldwide", "headquarter", "hq", "regional", "multi-country",
    "overseas", "foreign", "relocation", "anglophone",
]

# The candidate profile is the user's profile.md (facts only).
PROFILE_FACTS = jh_config.profile_text(_CFG)


# ── HELPERS ──────────────────────────────────────────────────────────────────
def score_job(title: str, company: str, jd: str) -> dict:
    if not ANTHROPIC_API_KEY:
        return {"score": 0, "strengths": "No API key", "gaps": "N/A"}
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = f"""You are a career advisor scoring job-fit for a candidate.

CANDIDATE PROFILE:
{PROFILE_FACTS}

JOB POSTING:
Title: {title}
Company: {company}
Description:
{jd[:3000]}

Score the fit from 0–100. Be strict and realistic.
Respond ONLY in this JSON format (no markdown, no commentary):
{{"score": <int>, "strengths": "<one sentence>", "gaps": "<one sentence>"}}"""
    last_err = ""
    for attempt in range(3):
        try:
            msg = client.messages.create(
                model=_CFG.get("scoring_model", "claude-haiku-4-5"),
                max_tokens=200,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw).strip()
            return json.loads(raw)
        except Exception as e:
            last_err = str(e)[:80]
            time.sleep(2 * (attempt + 1))  # 2s, 4s, 6s backoff on transient errors
    # ponytail: all retries failed. Do NOT return score 0 — that silently drops the
    # job below the 65 floor as if it were a bad match. Return it AT the floor with a
    # visible flag so it lands in the tracker for the user to rescore, instead of vanishing.
    return {"score": 65, "strengths": "SCORING FAILED — review/rescore manually", "gaps": last_err}


def is_international(title: str, company: str, jd: str) -> bool:
    text = (title + " " + company + " " + jd).lower()
    if any(c in text for c in INTL_COMPANY_LIST):
        return True
    if any(k in text for k in INTL_JD_KEYWORDS):
        return True
    latin_chars = sum(1 for c in jd if c.isascii() and c.isalpha())
    total_alpha = sum(1 for c in jd if c.isalpha())
    if total_alpha > 0 and latin_chars / total_alpha > 0.5:
        return True
    return False


def load_seen_jobs(wb: openpyxl.Workbook) -> set:
    seen = set()
    for sheet_name in ["Applications", "Candidates"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                company = str(row[1] or "").lower().strip()
                role = str(row[2] or "").lower().strip()
                if company and role:
                    seen.add((company, role))
    return seen


def ensure_candidates_sheet(wb: openpyxl.Workbook):
    if "Candidates" not in wb.sheetnames:
        ws = wb.create_sheet("Candidates")
        headers = [
            "Date Found", "Company", "Job Role", "Location",
            "International", "Match Score", "Strengths", "Key Gaps",
            "Source", "URL", "JD (truncated)", "Status"
        ]
        ws.append(headers)
        col_widths = [12, 22, 35, 14, 12, 12, 45, 45, 16, 40, 60, 12]
        for i, (cell, width) in enumerate(zip(ws[1], col_widths), 1):
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 28
    return wb["Candidates"]


def _no_formula(v):
    # ponytail: scraped text starting with '=' is stored by openpyxl as a LIVE
    # Excel formula (data_type 'f') — formula injection from a malicious job
    # posting. A leading space keeps it inert text. Only '=' matters for xlsx.
    return " " + v if isinstance(v, str) and v.startswith("=") else v


def write_candidate_row(ws, job: dict):
    score = job.get("score", 0)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if score >= 80:
        fill = PatternFill("solid", fgColor="C6EFCE")
    elif score >= 70:
        fill = PatternFill("solid", fgColor="FFEB9C")
    else:
        fill = PatternFill("solid", fgColor="EBEBEB")
    row_data = [
        job.get("date_found", ""),
        job.get("company", ""),
        job.get("title", ""),
        job.get("location", ""),
        "Yes" if job.get("international") else "",
        score,
        job.get("strengths", ""),
        job.get("gaps", ""),
        job.get("source", ""),
        job.get("url", ""),
        job.get("jd_snippet", "")[:500],
        "New",
    ]
    ws.append([_no_formula(x) for x in row_data])
    row_idx = ws.max_row
    for i, cell in enumerate(ws[row_idx], 1):
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=(i in [7, 8, 11]), vertical="top")
    ws.row_dimensions[row_idx].height = 22


# ── LINKEDIN SCRAPER ─────────────────────────────────────────────────────────
async def scrape_linkedin(searches: list, seen: set, max_per_search: int) -> list:
    """
    Opens a real Chrome window.
    Waits for you to log in if needed.
    Searches LinkedIn Jobs for each query + Israel location.
    Extracts job cards, navigates to detail panels for full JD.
    """
    all_jobs = []

    async with async_playwright() as p:
        print("\nOpening Chrome for LinkedIn...")
        print("If LinkedIn asks you to log in, do so now — the script will wait.")
        browser = await p.chromium.launch(
            channel="chrome",
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"]
        )
        ctx = await browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        await ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        page = await ctx.new_page()

        # Go to LinkedIn first — let user log in if not already
        print("\nNavigating to LinkedIn...")
        await page.goto("https://www.linkedin.com/jobs/", wait_until="load", timeout=30000)
        await asyncio.sleep(3)

        # Check for login wall
        current_url = page.url
        if "login" in current_url or "authwall" in current_url:
            print("\n>>> LinkedIn login required.")
            print(">>> Please log in now. The script will continue automatically once you're on the Jobs page.")
            # Wait up to 2 minutes for the user to log in
            try:
                await page.wait_for_url("**/jobs/**", timeout=120000)
                print(">>> Logged in. Continuing...")
            except Exception:
                print(">>> Timeout waiting for login. Proceeding anyway...")
        await asyncio.sleep(2)

        for query, geo_or_loc, location_label in searches:
            print(f"\n--- Searching LinkedIn: '{query}' ({location_label}) ---")
            encoded_q = query.replace(" ", "%20")
            if geo_or_loc.isdigit():
                location_param = f"&geoId={geo_or_loc}"
            else:
                location_param = f"&location={geo_or_loc.replace(' ', '%20')}"
            search_url = (
                f"https://www.linkedin.com/jobs/search/"
                f"?keywords={encoded_q}"
                f"{location_param}"
                f"&f_TPR=r86400"   # posted in last 24h — remove to get all recent
                f"&sortBy=R"        # relevance sort
            )
            await page.goto(search_url, wait_until="load", timeout=30000)
            await asyncio.sleep(4)

            # Scroll a couple of times to load more cards
            for _ in range(3):
                await page.keyboard.press("End")
                await asyncio.sleep(1.5)

            # LinkedIn job cards
            cards = await page.query_selector_all(".job-search-card, .jobs-search__results-list li, .base-card")
            print(f"  Found {len(cards)} job cards")

            stubs = []
            for card in cards:
                try:
                    title_el = await card.query_selector(".base-search-card__title, h3, .job-search-card__title")
                    company_el = await card.query_selector(".base-search-card__subtitle, h4, .job-search-card__company-name")
                    location_el = await card.query_selector(".job-search-card__location, .base-search-card__metadata")
                    link_el = await card.query_selector("a[href*='/jobs/']")

                    title = (await title_el.inner_text()).strip() if title_el else ""
                    company = (await company_el.inner_text()).strip() if company_el else ""
                    location = (await location_el.inner_text()).strip() if location_el else ""
                    href = (await link_el.get_attribute("href")) if link_el else ""

                    if not title or not href:
                        continue

                    # Strip tracking params from URL
                    clean_url = href.split("?")[0] if "?" in href else href
                    if not clean_url.startswith("http"):
                        clean_url = "https://www.linkedin.com" + clean_url

                    # Skip non-managerial titles
                    mgr_kw = ["manager", "director", "head of", "vp ", "מנהל", "ראש", "senior", "pmo"]
                    if not any(k in title.lower() for k in mgr_kw):
                        continue

                    key = (company.lower().strip(), title.lower().strip())
                    if key in seen:
                        continue

                    stubs.append({
                        "title": title, "company": company,
                        "location": location, "url": clean_url
                    })
                    seen.add(key)
                    if len(stubs) >= max_per_search:
                        break
                except Exception:
                    continue

            print(f"  {len(stubs)} new managerial jobs to fetch")

            # PASS 2: Click each card and extract the JD from the side panel
            for i, job in enumerate(stubs):
                print(f"  [{i+1}/{len(stubs)}] {job['title'][:50]} @ {job['company'][:25]}")
                try:
                    await page.goto(job["url"], wait_until="load", timeout=20000)
                    await asyncio.sleep(SLEEP_BETWEEN_JOBS)

                    # LinkedIn shows JD in a details panel
                    jd = ""
                    for jd_sel in [
                        ".show-more-less-html__markup",
                        ".jobs-description__content",
                        ".jobs-description",
                        ".description__text",
                        "article",
                        "main",
                    ]:
                        el = await page.query_selector(jd_sel)
                        if el:
                            jd = (await el.inner_text())[:3000]
                            if len(jd) > 80:
                                break
                    if not jd:
                        jd = (await page.inner_text("body"))[:2000]

                    job["jd"] = jd
                    job["source"] = f"LinkedIn: {query} ({location_label})"
                except Exception as e:
                    job["jd"] = ""
                    job["source"] = f"LinkedIn: {query} ({location_label})"
                    print(f"    Error: {e}")

            all_jobs.extend(stubs)

        print("\nAll LinkedIn searches done. Closing browser in 3s...")
        await asyncio.sleep(3)
        await browser.close()

    return all_jobs


# ── MAIN ─────────────────────────────────────────────────────────────────────
async def main():
    today = datetime.date.today().strftime("%d/%m/%Y")
    print(f"\n{'='*60}")
    print(f"LinkedIn Manual Crawler — {today}")
    print(f"{'='*60}")

    if not ANTHROPIC_API_KEY:
        print("ERROR: ANTHROPIC_API_KEY not set. Open a new terminal and try again.")
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    seen = load_seen_jobs(wb)
    ws_candidates = ensure_candidates_sheet(wb)
    print(f"Loaded tracker. {len(seen)} jobs already seen (dedup active).")

    jobs = await scrape_linkedin(LINKEDIN_SEARCHES, seen, MAX_JOBS_PER_SEARCH)
    print(f"\nTotal jobs collected: {len(jobs)}")

    if not jobs:
        print("Nothing to score. Done.")
        return

    print(f"\nScoring {len(jobs)} jobs with Claude...")
    new_count = 0
    for job in jobs:
        jd = job.get("jd", "")
        print(f"  Scoring: {job['title'][:50]} @ {job['company'][:25]}...")
        result = score_job(job["title"], job["company"], jd)
        score = result.get("score", 0)
        intl = is_international(job["title"], job["company"], jd)
        print(f"  Score: {score} | International: {intl}")

        entry = {
            "date_found": today,
            "company": job["company"],
            "title": job["title"],
            "location": job.get("location", ""),
            "international": intl,
            "score": score,
            "strengths": result.get("strengths", ""),
            "gaps": result.get("gaps", ""),
            "source": job.get("source", "LinkedIn"),
            "url": job.get("url", ""),
            "jd_snippet": jd[:500],
        }
        skip, reason = should_skip(score, entry["location"], job["title"], jd, is_international=intl)
        if skip:
            print(f"  Skipped ({reason})")
            continue
        write_candidate_row(ws_candidates, entry)
        new_count += 1
        time.sleep(0.3)

    # NOTE: No sort step — existing rows are never rewritten.
    # New rows are appended at the bottom. User edits (e.g. Status changes) are preserved.

    wb.save(TRACKER_PATH)
    print(f"\n{'='*60}")
    print(f"Done. Added {new_count} new candidates.")
    print(f"File saved: {TRACKER_PATH}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
