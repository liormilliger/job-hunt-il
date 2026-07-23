# -*- coding: utf-8 -*-
"""job-hunt-il first-run setup.

Usage:  python setup.py

1. Checks dependencies and tells you exactly what's missing.
2. Creates config.json from config.example.json (if absent).
3. Creates the workdir, the Excel tracker, and scaffolds profile.md +
   positioning.md from templates/.

After this script, open a Claude session and ask it to interview you to fill
profile.md and positioning.md — that conversation is where the pipeline gets
its brain.
"""
import json
import os
import shutil
import subprocess
import sys

sys.stdout.reconfigure(encoding="utf-8")
SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def check(name, fn):
    try:
        ok, note = fn()
    except Exception as e:
        ok, note = False, str(e)[:60]
    print(f"  [{'OK' if ok else 'MISSING'}] {name}" + (f" — {note}" if note else ""))
    return ok


def dep_checks():
    print("Dependency check:")
    results = []

    def py_pkg(mod, pipname=None):
        def f():
            __import__(mod)
            return True, ""
        return check(f"python: {pipname or mod}", f)

    results.append(py_pkg("openpyxl"))
    results.append(py_pkg("anthropic"))
    results.append(py_pkg("docx", "python-docx"))
    results.append(py_pkg("playwright"))
    results.append(py_pkg("docx2pdf"))

    def node():
        v = subprocess.run(["node", "--version"], capture_output=True, text=True)
        return v.returncode == 0, v.stdout.strip()
    results.append(check("node (English CV renderer)", node))

    def word():
        if os.name != "nt":
            return False, "Windows + MS Word needed for PDF export"
        try:
            import winreg
            winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, "Word.Application")
            return True, ""
        except OSError:
            return False, "MS Word not found (PDF export will fail)"
    results.append(check("Microsoft Word", word))

    def chromium():
        r = subprocess.run([sys.executable, "-m", "playwright", "install",
                            "--dry-run", "chromium"], capture_output=True, text=True)
        return r.returncode == 0, "run: python -m playwright install chromium"
    results.append(check("playwright chromium", chromium))

    if not all(results):
        print("\nInstall the missing pieces:")
        print("  pip install openpyxl anthropic python-docx playwright docx2pdf")
        print("  python -m playwright install chromium")
        print("  (cd scripts/cv_render && npm install)   # English CV renderer")
    return all(results)


def make_tracker(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    def sheet(name, headers, widths):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for i, (cell, w) in enumerate(zip(ws[1], widths), 1):
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28
        return ws

    sheet("Applications",
          ["Date", "Company", "Role", "Description", "Match Score",
           "CV", "Cover Letter", "Status", "Notes"],
          [12, 22, 32, 50, 12, 30, 30, 14, 40])
    sheet("Candidates",
          ["Date Found", "Company", "Job Role", "Location", "International",
           "Match Score", "Strengths", "Key Gaps", "Source", "URL",
           "JD (truncated)", "Status", "", "CV Folder", "Gen CV"],
          [12, 22, 35, 14, 12, 12, 45, 45, 16, 40, 60, 12, 4, 34, 8])
    del wb["Sheet"]
    wb.save(path)
    print(f"  created tracker: {path}")


def main():
    print("=== job-hunt-il setup ===\n")
    deps_ok = dep_checks()

    cfg_path = os.path.join(SKILL_ROOT, "config.json")
    if not os.path.exists(cfg_path):
        shutil.copy(os.path.join(SKILL_ROOT, "config.example.json"), cfg_path)
        print(f"\ncreated {cfg_path} — EDIT IT NOW (workdir, name, search titles), "
              "then run setup again to scaffold the workdir.")
        return
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)

    wd = cfg.get("workdir", "")
    if not wd or "you" in wd.lower().replace("your", ""):
        print("\nconfig.json still has the example workdir — edit it first.")
        return
    os.makedirs(wd, exist_ok=True)

    tracker = os.path.join(wd, cfg.get("tracker", "Job_Tracker.xlsx"))
    if not os.path.exists(tracker):
        make_tracker(tracker)
    else:
        print(f"  tracker exists: {tracker}")

    for key, tmpl in (("profile_md", "profile_template.md"),
                      ("positioning_md", "positioning_template.md")):
        dst = os.path.join(wd, cfg.get(key, tmpl.replace("_template", "")))
        if not os.path.exists(dst):
            shutil.copy(os.path.join(SKILL_ROOT, "templates", tmpl), dst)
            print(f"  scaffolded {dst}")

    print("\nSetup done." if deps_ok else "\nSetup done, but fix the MISSING deps above.")
    print("Next: ask Claude to interview you and fill profile.md + positioning.md.")


if __name__ == "__main__":
    main()
