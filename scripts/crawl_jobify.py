"""
Jobify360 Crawler — job-hunt-il
===================================================
Scrapes https://jobify360.co.il/myjobs (your personalised role list).
Opens a real Chrome window — log in if prompted, then the script runs automatically.

Run:  python crawl_jobify.py
"""

import sys
import os
import asyncio
import json
import re
import time
import datetime
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.async_api import async_playwright

from crawler_filters import should_skip

sys.stdout.reconfigure(encoding="utf-8")

# ── CONFIG ───────────────────────────────────────────────────────────────────
import jh_config
_CFG = jh_config.load()
TRACKER_PATH = _CFG["tracker"]
BASE_URL = "https://jobify360.co.il"
MYJOBS_URL = f"{BASE_URL}/myjobs"

ANTHROPIC_API_KEY = jh_config.api_key(_CFG)

SCORE_THRESHOLD = 70
SLEEP_BETWEEN_PAGES = 2.0

# Roles to scrape — matched against the Hebrew/English tab titles on /myjobs.
# Comes from config "role_keywords"; the fallback list covers common
# operations / supply-chain / project keywords in both languages.
TARGET_ROLE_KEYWORDS = _CFG.get("role_keywords") or [
    "תפעול", "operations",
    "שרשרת", "supply",
    "לוגיסטיקה", "logistics",
    "פרויקט", "project", "pmo", "program",
    "תכנון", "planning",
    "מחסן", "warehouse",
    "הפצה", "distribution",
]

INTL_COMPANY_LIST = [
    "nike", "adidas", "l'oreal", "loreal", "henkel", "p&g", "procter", "unilever",
    "amazon", "dhl", "maersk", "ups", "fedex", "h&m", "ikea", "zara", "inditex",
    "nestle", "abbott", "johnson", "pfizer", "novartis", "roche", "siemens",
    "microsoft", "google", "apple", "intel", "applied materials", "kla", "nova",
    "elta", "rafael", "elbit", "checkpoint", "amdocs", "nice", "radcom",
    "teva", "given imaging", "orbotech", "strauss", "osem", "tnuva",
]
INTL_JD_KEYWORDS = [
    "emea", "global", "international", "cross-border", "import", "export",
    "fluent english", "english required", "english speaker", "multinational",
    "worldwide", "headquarter", "hq", "regional", "multi-country",
    "overseas", "foreign", "relocation", "anglophone",
]

# The candidate profile is the user's profile.md (facts only).
PROFILE_FACTS = jh_config.profile_text(_CFG)


# ── HELPERS ──────────────────────────────────────────────────────────────────
def score_job(title: str, company: str, jd: str) -> dict:
    if not ANTHROPIC_API_KEY:
        return {"score": 0, "strengths": "No API key", "gaps": "N/A"}
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = f"""You are a career advisor scoring job-fit for a candidate.

CANDIDATE PROFILE:
{PROFILE_FACTS}

JOB POSTING:
Title: {title}
Company: {company}
Description:
{jd[:3000]}

Score the fit from 0–100. Be strict and realistic.
Respond ONLY in this JSON format (no markdown, no commentary):
{{"score": <int>, "strengths": "<one sentence>", "gaps": "<one sentence>"}}"""
    last_err = ""
    for attempt in range(3):
        try:
            msg = client.messages.create(
                model=_CFG.get("scoring_model", "claude-haiku-4-5"),
                max_tokens=200,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw).strip()
            return json.loads(raw)
        except Exception as e:
            last_err = str(e)[:80]
            time.sleep(2 * (attempt + 1))  # 2s, 4s, 6s backoff on transient errors
    # ponytail: all retries failed. Do NOT return score 0 — that silently drops the
    # job below the 65 floor as if it were a bad match. Return it AT the floor with a
    # visible flag so it lands in the tracker for the user to rescore, instead of vanishing.
    return {"score": 65, "strengths": "SCORING FAILED — review/rescore manually", "gaps": last_err}


def is_international(title: str, company: str, jd: str) -> bool:
    text = (title + " " + company + " " + jd).lower()
    if any(c in text for c in INTL_COMPANY_LIST):
        return True
    if any(k in text for k in INTL_JD_KEYWORDS):
        return True
    latin_chars = sum(1 for c in jd if c.isascii() and c.isalpha())
    total_alpha = sum(1 for c in jd if c.isalpha())
    if total_alpha > 0 and latin_chars / total_alpha > 0.5:
        return True
    return False


def load_seen_jobs(wb: openpyxl.Workbook) -> set:
    seen = set()
    for sheet_name in ["Applications", "Candidates"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                company = str(row[1] or "").lower().strip()
                role = str(row[2] or "").lower().strip()
                if company and role:
                    seen.add((company, role))
    return seen


def ensure_candidates_sheet(wb: openpyxl.Workbook):
    if "Candidates" not in wb.sheetnames:
        ws = wb.create_sheet("Candidates")
        headers = [
            "Date Found", "Company", "Job Role", "Location",
            "International", "Match Score", "Strengths", "Key Gaps",
            "Source", "URL", "JD (truncated)", "Status"
        ]
        ws.append(headers)
        col_widths = [12, 22, 35, 14, 12, 12, 45, 45, 16, 40, 60, 12]
        for i, (cell, width) in enumerate(zip(ws[1], col_widths), 1):
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 28
    return wb["Candidates"]


def write_candidate_row(ws, job: dict):
    score = job.get("score", 0)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if score >= 80:
        fill = PatternFill("solid", fgColor="C6EFCE")
    elif score >= 70:
        fill = PatternFill("solid", fgColor="FFEB9C")
    else:
        fill = PatternFill("solid", fgColor="EBEBEB")
    row_data = [
        job.get("date_found", ""),
        job.get("company", ""),
        job.get("title", ""),
        job.get("location", ""),
        "Yes" if job.get("international") else "",
        score,
        job.get("strengths", ""),
        job.get("gaps", ""),
        job.get("source", ""),
        job.get("url", ""),
        job.get("jd_snippet", "")[:500],
        "New",
    ]
    ws.append(row_data)
    row_idx = ws.max_row
    for i, cell in enumerate(ws[row_idx], 1):
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=(i in [7, 8, 11]), vertical="top")
    ws.row_dimensions[row_idx].height = 22


# ── SCRAPER ───────────────────────────────────────────────────────────────────
def _is_target_role(tab_title: str) -> bool:
    t = tab_title.lower()
    return any(kw in t for kw in TARGET_ROLE_KEYWORDS)


async def extract_cards_from_role_page(page) -> list:
    """Extract job stubs from a /myjob-roles/... page. Tries three selector strategies."""
    # Scroll a few times to trigger lazy-load
    for _ in range(3):
        await page.keyboard.press("End")
        await asyncio.sleep(1)

    # Strategy 1: original selectors (Jobify360 v1 DOM)
    stubs = await page.evaluate("""
        () => {
            const wrappers = document.querySelectorAll('.w-100.d-flex.flex-column');
            const jobs = [];
            for (const w of wrappers) {
                const headerStart = w.querySelector('.card-header-start');
                if (!headerStart) continue;
                const lines = headerStart.innerText.trim().split('\\n')
                                .map(l => l.trim()).filter(Boolean);
                const title   = lines[0] || '';
                const company = lines[1] || '';
                const snippet = (w.querySelector('.card-body') || {innerText:''})
                                  .innerText.trim().slice(0, 400);
                const linkEl  = w.querySelector('a[href*="/jobs/"]');
                const href    = linkEl ? linkEl.href : '';
                if (title && href) jobs.push({title, company, snippet, href});
            }
            return jobs;
        }
    """)
    if stubs:
        print(f"    [selector=strategy1] {len(stubs)} cards")
        return stubs

    # Strategy 2: find all /jobs/ links and walk up to nearest card container
    stubs = await page.evaluate("""
        () => {
            const jobs = [];
            const seen = new Set();
            const links = document.querySelectorAll('a[href*="/jobs/"]');
            for (const link of links) {
                const href = link.href;
                if (!href || seen.has(href)) continue;
                seen.add(href);
                // Walk up to find a reasonable container (up to 6 levels)
                let container = link.parentElement;
                for (let i = 0; i < 5; i++) {
                    if (!container) break;
                    const rect = container.getBoundingClientRect();
                    if (rect.height > 60) break;
                    container = container.parentElement;
                }
                const text = (container || link.parentElement || link).innerText || '';
                const lines = text.trim().split('\\n').map(l => l.trim()).filter(Boolean);
                const title   = lines[0] || link.innerText.trim() || '';
                const company = lines[1] || '';
                const snippet = lines.slice(2).join(' ').slice(0, 400);
                if (title) jobs.push({title, company, snippet, href});
            }
            return jobs;
        }
    """)
    if stubs:
        print(f"    [selector=strategy2/job-links] {len(stubs)} cards")
        return stubs

    # Strategy 3: any .card / article / element with "job" in class that has a link
    stubs = await page.evaluate("""
        () => {
            const jobs = [];
            const seen = new Set();
            const cards = document.querySelectorAll(
                '.card, article, [class*="job-card"], [class*="position"], [class*="vacancy"]'
            );
            for (const card of cards) {
                const linkEl = card.querySelector('a[href]');
                if (!linkEl) continue;
                const href = linkEl.href || '';
                if (seen.has(href)) continue;
                seen.add(href);
                const text = card.innerText.trim();
                const lines = text.split('\\n').map(l => l.trim()).filter(Boolean);
                const title   = lines[0] || '';
                const company = lines[1] || '';
                const snippet = lines.slice(2).join(' ').slice(0, 400);
                if (title && href) jobs.push({title, company, snippet, href});
            }
            return jobs;
        }
    """)
    if stubs:
        print(f"    [selector=strategy3/cards] {len(stubs)} cards")
        return stubs

    # Debug dump so we know what the page actually looks like
    url = page.url
    body_sample = (await page.inner_text("body"))[:400].replace("\n", " ")
    print(f"    [DEBUG] 0 cards found. URL={url}")
    print(f"    [DEBUG] Page sample: {body_sample}")
    return []


async def fetch_full_jd(page, job_url: str) -> str:
    """Navigate to the job detail page and return the full JD text.

    Selectors verified against the live Jobify360 site on 10/06/2026:
    the JD sits in the largest .white-border-box card inside .content-text-main.
    The old guessed selectors (.job-description etc.) matched nothing, so every
    job fell through to a whole-page grab full of site boilerplate.
    """
    try:
        await page.goto(job_url, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(SLEEP_BETWEEN_PAGES)

        # 1. The JD card: longest white-border-box that isn't the AI-tips
        #    widget or the match-score gauge
        boxes = await page.query_selector_all(".content-text-main .white-border-box")
        candidates = []
        for box in boxes:
            text = (await box.inner_text()).strip()
            if len(text) > 120 and "כל מה שצריך לדעת" not in text and "ציון התאמה" not in text:
                candidates.append(text)
        if candidates:
            return max(candidates, key=len)[:3000]

        # 2. Fallback: the whole content area, minus the AI-tips widget header
        el = await page.query_selector(".content-text-main")
        if el:
            text = (await el.inner_text()).strip()
            marker = "הכנה לראיון עבודה"  # last line of the AI-tips widget
            if marker in text:
                text = text.split(marker, 1)[1].strip()
            if len(text) > 80:
                return text[:3000]

        # 3. Last resort: page body, hard-cut before the site footer so
        #    boilerplate can never reach the tracker again
        body = await page.inner_text("body")
        for stop in ("תנאי שימוש", "מדיניות פרטיות", "מי אנחנו"):
            idx = body.find(stop)
            if idx > 200:
                body = body[:idx]
                break
        return body[:2000]
    except Exception as e:
        return f"[Error: {e}]"


async def scrape_jobify360(seen: set) -> list:
    all_stubs = []

    async with async_playwright() as p:
        print("\nOpening Chrome for Jobify360...")
        browser = await p.chromium.launch(
            channel="chrome",
            headless=False,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"]
        )
        ctx = await browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )
        await ctx.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = await ctx.new_page()

        # ── Step 1: load /myjobs and handle login ──────────────────────────
        print(f"Navigating to {MYJOBS_URL} ...")
        await page.goto(MYJOBS_URL, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(3)

        # If not logged in, the page won't show "שלום" — wait for it
        for attempt in range(24):   # up to 2 minutes
            body_text = await page.inner_text("body")
            if "שלום" in body_text or "myjob-roles" in page.url:
                break
            if attempt == 0:
                print(">>> Jobify360 login required.")
                print(">>> Please log in in the Chrome window. The script will continue automatically.")
            await asyncio.sleep(5)
        else:
            print(">>> Timed out waiting for login. Continuing anyway...")

        await asyncio.sleep(2)

        # Jobify occasionally throws announcement modals (e.g. the
        # "What's New" popup, seen live 22/07/2026) that sit over the page and
        # intercept every click until dismissed. Kill any open modal +
        # backdrop before interacting; harmless when none is present.
        async def dismiss_modals():
            try:
                await page.keyboard.press("Escape")
                await page.evaluate(
                    "() => {"
                    "  document.querySelectorAll('.modal.show, .modal-backdrop')"
                    "    .forEach(e => e.remove());"
                    "  document.body.classList.remove('modal-open');"
                    "  document.body.style.overflow = '';"
                    "}"
                )
            except Exception:
                pass
        await dismiss_modals()

        # ── Step 2: get the tab list from /myjobs ─────────────────────────
        # Try multiple selectors in case Jobify360 updated their markup
        tab_titles = await page.evaluate("""
            () => {
                const selectors = [
                    'a.tab',
                    'a.tab.w-full',
                    '[class*="tab"] a',
                    'a[class*="tab"]',
                    '#accordionMyjob a',
                    '.accordion a',
                    '.sidebar a[href*="myjob"]',
                    'nav a[href*="myjob"]',
                ];
                for (const sel of selectors) {
                    const els = Array.from(document.querySelectorAll(sel));
                    if (els.length > 0) {
                        console.log('Tab selector matched: ' + sel + ' (' + els.length + ')');
                        return els.map(a => ({
                            title: a.innerText.trim().replace(/\\s+/g, ' '),
                            href: a.href || ''
                        }));
                    }
                }
                // Last resort: any link that looks like a category/role filter
                const all = Array.from(document.querySelectorAll('a[href]'))
                    .filter(a => a.href.includes('myjob') || a.href.includes('role') || a.href.includes('category'));
                return all.map(a => ({
                    title: a.innerText.trim().replace(/\\s+/g, ' '),
                    href: a.href || ''
                }));
            }
        """)
        print(f"\nFound {len(tab_titles)} role tabs on /myjobs:")
        if not tab_titles:
            body_sample = (await page.inner_text("body"))[:400].replace("\n", " ")
            print(f"  [DEBUG] No tabs found. URL={page.url}")
            print(f"  [DEBUG] Page sample: {body_sample}")
        for t in tab_titles:
            flag = "  [TARGET]" if _is_target_role(t["title"]) else ""
            print(f"  {t['title'][:60]}{flag}")

        # ── Step 3: click each target tab, scrape, come back ──────────────
        target_count = sum(1 for t in tab_titles if _is_target_role(t["title"]))
        print(f"\nScraping {target_count} target role(s)...\n")

        scraped_role_urls = set()   # avoid visiting same role page twice

        for idx, tab_info in enumerate(tab_titles):
            if not _is_target_role(tab_info["title"]):
                continue

            # If the tab carries an href, navigate directly — faster and more reliable
            tab_href = tab_info.get("href", "")
            if tab_href and tab_href.startswith("http") and "myjob" in tab_href:
                print(f"Navigating directly: {tab_info['title'][:60]}")
                await page.goto(tab_href, wait_until="domcontentloaded", timeout=20000)
                await asyncio.sleep(2)
                await dismiss_modals()
            else:
                # Fall back to re-loading /myjobs and clicking by index
                if page.url.rstrip("/") != MYJOBS_URL.rstrip("/"):
                    await page.goto(MYJOBS_URL, wait_until="domcontentloaded", timeout=20000)
                    await asyncio.sleep(2)

                # Re-query tabs after reload using the same multi-selector approach
                tabs = await page.query_selector_all(
                    "a.tab, a.tab.w-full, a[class*='tab'], #accordionMyjob a"
                )
                if idx >= len(tabs):
                    print(f"  [SKIP] Tab index {idx} out of range after reload")
                    continue

                print(f"Clicking: {tab_info['title'][:60]}")
                await dismiss_modals()
                await tabs[idx].click()
                # Wait for either a myjob-roles URL or just a network idle
                try:
                    await page.wait_for_url("**/myjob-roles/**", timeout=8000)
                except Exception:
                    # URL may not change (SPA rendering) — wait for content instead
                    try:
                        await page.wait_for_load_state("networkidle", timeout=6000)
                    except Exception:
                        pass
                await asyncio.sleep(2)

            role_url = page.url
            if role_url in scraped_role_urls:
                print(f"  [SKIP] Already scraped {role_url}")
                continue
            scraped_role_urls.add(role_url)

            print(f"  Role page: {role_url}")
            stubs = await extract_cards_from_role_page(page)
            print(f"  Found {len(stubs)} job cards")

            for stub in stubs:
                key = (stub["company"].lower().strip(), stub["title"].lower().strip())
                if key in seen:
                    print(f"    [SKIP dedup] {stub['title'][:40]} @ {stub['company'][:25]}")
                    continue
                seen.add(key)
                stub["role_tab"] = tab_info["title"].split("\n")[0].strip()
                all_stubs.append(stub)

        # ── Step 4: fetch full JDs ─────────────────────────────────────────
        print(f"\nFetching full JDs for {len(all_stubs)} jobs...")
        for i, stub in enumerate(all_stubs):
            print(f"  [{i+1}/{len(all_stubs)}] {stub['title'][:50]} @ {stub['company'][:25]}")
            stub["jd"] = await fetch_full_jd(page, stub["href"])

        print("\nAll done. Closing browser in 3s...")
        await asyncio.sleep(3)
        await browser.close()

    return all_stubs


# ── MAIN ─────────────────────────────────────────────────────────────────────
async def main():
    today = datetime.date.today().strftime("%d/%m/%Y")
    print(f"\n{'='*60}")
    print(f"Jobify360 Manual Crawler — {today}")
    print(f"{'='*60}")

    if not ANTHROPIC_API_KEY:
        print("ERROR: ANTHROPIC_API_KEY not set. Open a new terminal and try again.")
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    seen = load_seen_jobs(wb)
    ws_candidates = ensure_candidates_sheet(wb)
    print(f"Loaded tracker. {len(seen)} jobs already seen (dedup active).")

    stubs = await scrape_jobify360(seen)
    print(f"\nTotal new jobs collected: {len(stubs)}")

    if not stubs:
        print("Nothing to score. Done.")
        return

    print(f"\nScoring {len(stubs)} jobs with Claude...")
    new_count = 0
    for stub in stubs:
        jd = stub.get("jd", stub.get("snippet", ""))
        title = stub["title"]
        company = stub["company"]
        print(f"  Scoring: {title[:50]} @ {company[:25]}...")

        result = score_job(title, company, jd)
        score = result.get("score", 0)
        intl = is_international(title, company, jd)
        print(f"  Score: {score} | International: {intl}")

        entry = {
            "date_found": today,
            "company": company,
            "title": title,
            "location": "",
            "international": intl,
            "score": score,
            "strengths": result.get("strengths", ""),
            "gaps": result.get("gaps", ""),
            "source": f"Jobify360: {stub.get('role_tab', '')}",
            "url": stub.get("href", ""),
            "jd_snippet": jd[:500],
        }
        skip, reason = should_skip(score, entry["location"], title, jd, is_international=intl)
        if skip:
            print(f"  Skipped ({reason})")
            continue
        write_candidate_row(ws_candidates, entry)
        new_count += 1
        time.sleep(0.3)

    # NOTE: No sort step — existing rows are never rewritten.
    # New rows are appended at the bottom. User edits (e.g. Status changes) are preserved.

    wb.save(TRACKER_PATH)
    print(f"\n{'='*60}")
    print(f"Done. Added {new_count} new candidates to tracker.")
    print(f"File: {TRACKER_PATH}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
