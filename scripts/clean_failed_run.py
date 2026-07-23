# -*- coding: utf-8 -*-
"""Remove a failed crawl batch from the Candidates sheet.

Usage:  python clean_failed_run.py DD/MM/YYYY

A row is deleted ONLY if BOTH the date matches AND it carries the full
SCORING-FAILED signature (score == min_score, 'FAILED' in Strengths, status
'New', no CV Folder). Rows the user has touched — status changed, CV
generated — can never be caught, even on the right date.
"""
import sys
import openpyxl
import jh_config

sys.stdout.reconfigure(encoding="utf-8")


def main():
    if len(sys.argv) != 2:
        sys.exit("Usage: python clean_failed_run.py DD/MM/YYYY")
    target_date = sys.argv[1]

    cfg = jh_config.load()
    wb = openpyxl.load_workbook(cfg["tracker"])
    ws = wb["Candidates"]
    hdr = [c.value for c in ws[1]]

    def ci(name):
        for i, h in enumerate(hdr, 1):
            if h and name.lower() in str(h).lower():
                return i
        return None

    C_SCORE, C_STR, C_STAT, C_CVF = (ci("Match Score"), ci("Strengths"),
                                     ci("Status"), ci("CV Folder"))
    min_score = cfg.get("min_score", 65)

    to_delete, kept = [], 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) != target_date:
            continue
        failed = (ws.cell(r, C_SCORE).value == min_score
                  and "FAILED" in str(ws.cell(r, C_STR).value or "").upper()
                  and str(ws.cell(r, C_STAT).value or "") == "New"
                  and not ws.cell(r, C_CVF).value)
        if failed:
            to_delete.append(r)
        else:
            kept += 1

    print(f"matched {len(to_delete)} failed rows on {target_date}; "
          f"kept {kept} non-matching rows from that date")
    for r in sorted(to_delete, reverse=True):   # bottom-up: indexes stay valid
        ws.delete_rows(r, 1)
    wb.save(cfg["tracker"])
    print(f"deleted {len(to_delete)}; Candidates now has {ws.max_row} rows. SAVED")


if __name__ == "__main__":
    main()
